import io
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================
# PDF EXPORTS
# =============================================

def generate_pl_pdf(report, branch_name, period):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Colors
    dark = colors.HexColor('#1e293b')
    green = colors.HexColor('#16a34a')
    red = colors.HexColor('#dc2626')
    indigo = colors.HexColor('#4f46e5')
    light_gray = colors.HexColor('#f8fafc')
    mid_gray = colors.HexColor('#e2e8f0')

    title_style = ParagraphStyle('title', fontSize=20, textColor=dark, fontName='Helvetica-Bold', spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=2)
    section_style = ParagraphStyle('section', fontSize=12, textColor=indigo, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6)

    # Header
    elements.append(Paragraph('Blacphics', title_style))
    elements.append(Paragraph(f'Profit & Loss Report — {branch_name}', sub_style))
    elements.append(Paragraph(f'Period: {period.title()} | Generated: {date.today().strftime("%d %b %Y")}', sub_style))
    elements.append(HRFlowable(width='100%', thickness=1, color=indigo, spaceAfter=12))

    def money(v):
        try:
            val = float(v or 0)
            return f"${val:,.2f}"
        except:
            return "$0.00"

    def color_val(v):
        try:
            return green if float(v or 0) >= 0 else red
        except:
            return dark

    s = report.get('sales', {})
    e = report.get('expenses', {})

    # Revenue section
    elements.append(Paragraph('Revenue', section_style))
    rev_data = [
        ['', 'Amount'],
        ['Gross Sales', money(s.get('gross_sales'))],
        ['Discounts', f"−{money(s.get('discounts'))}"],
        ['Net Sales', money(s.get('net_sales'))],
        ['Manual Revenue', money(s.get('manual_revenue'))],
        ['Total Revenue', money(s.get('total_revenue'))],
    ]
    rev_table = Table(rev_data, colWidths=[12*cm, 4*cm])
    rev_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), indigo),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, light_gray]),
        ('BACKGROUND', (0,-1), (-1,-1), mid_gray),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.3, mid_gray),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(rev_table)
    elements.append(Spacer(1, 0.4*cm))

    # COGS & Profit section
    elements.append(Paragraph('Cost of Goods & Profit', section_style))
    cogs_data = [
        ['', 'Amount'],
        ['COGS', money(report.get('cogs'))],
        ['Gross Profit', money(report.get('gross_profit'))],
        ['Gross Margin', f"{report.get('gross_margin_pct', 0)}%"],
        ['Manual Expenses', money(e.get('manual'))],
        ['Supplier Payments', money(e.get('supplier_payments'))],
        ['Total Expenses', money(e.get('total'))],
        ['Net Profit', money(report.get('net_profit'))],
        ['Net Margin', f"{report.get('net_margin_pct', 0)}%"],
    ]
    cogs_table = Table(cogs_data, colWidths=[12*cm, 4*cm])
    cogs_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), indigo),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-3), [colors.white, light_gray]),
        ('BACKGROUND', (0,-2), (-1,-1), mid_gray),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.3, mid_gray),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(cogs_table)
    elements.append(Spacer(1, 0.4*cm))

    # Accounts Receivable
    elements.append(Paragraph('Accounts Receivable', section_style))
    ar_data = [
        ['', 'Amount'],
        ['Total Collected', money(s.get('total_collected'))],
        ['Outstanding (AR)', money(s.get('accounts_receivable'))],
        ['Supplier Outstanding (AP)', money(e.get('supplier_outstanding'))],
    ]
    ar_table = Table(ar_data, colWidths=[12*cm, 4*cm])
    ar_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), indigo),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_gray]),
        ('GRID', (0,0), (-1,-1), 0.3, mid_gray),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(ar_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_orders_pdf(orders, branch_name):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []
    indigo = colors.HexColor('#4f46e5')
    light_gray = colors.HexColor('#f8fafc')
    mid_gray = colors.HexColor('#e2e8f0')
    dark = colors.HexColor('#1e293b')

    title_style = ParagraphStyle('title', fontSize=18, textColor=dark, fontName='Helvetica-Bold', spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=2)

    elements.append(Paragraph('Blacphics', title_style))
    elements.append(Paragraph(f'Orders Report — {branch_name}', sub_style))
    elements.append(Paragraph(f'Generated: {date.today().strftime("%d %b %Y")}', sub_style))
    elements.append(HRFlowable(width='100%', thickness=1, color=indigo, spaceAfter=10))

    headers = ['Order #', 'Customer', 'Type', 'Status', 'Payment', 'Total', 'Paid', 'Balance']
    data = [headers]

    for o in orders:
        customer = o.get('customer_name') or 'Walk-in'
        data.append([
            o.get('order_number', ''),
            customer[:20],
            o.get('transaction_type', '').replace('_', ' ').title(),
            o.get('status', '').title(),
            o.get('payment_status', '').title(),
            f"${float(o.get('total_amount', 0)):,.2f}",
            f"${float(o.get('amount_paid', 0)):,.2f}",
            f"${float(o.get('balance_due', 0)):,.2f}",
        ])

    col_widths = [2.5*cm, 3.5*cm, 2.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), indigo),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_gray]),
        ('GRID', (0,0), (-1,-1), 0.3, mid_gray),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('ALIGN', (5,0), (-1,-1), 'RIGHT'),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


# =============================================
# EXCEL EXPORTS
# =============================================

def generate_pl_excel(report, branch_name, period):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'P&L Report'

    indigo = 'FF4F46E5'
    light = 'FFF8FAFC'
    mid = 'FFE2E8F0'
    green_c = 'FF16A34A'
    red_c = 'FFDC2626'

    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='FF1E293B')
    section_font = Font(bold=True, size=11, color='FF4F46E5')
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)

    thin = Side(style='thin', color='FFE2E8F0')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def money(v):
        try: return float(v or 0)
        except: return 0.0

    def set_row(ws, row, col1, col2, font=None, fill=None, num_format='#,##0.00'):
        c1 = ws.cell(row=row, column=1, value=col1)
        c2 = ws.cell(row=row, column=2, value=col2)
        c1.border = thin_border
        c2.border = thin_border
        c2.number_format = f'$#,##0.00'
        c2.alignment = Alignment(horizontal='right')
        if font:
            c1.font = font
            c2.font = font
        if fill:
            c1.fill = fill
            c2.fill = fill
        return row + 1

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Blacphics — Profit & Loss Report'
    ws['A1'].font = title_font
    ws.merge_cells('A2:B2')
    ws['A2'] = f'Branch: {branch_name} | Period: {period.title()} | Generated: {date.today().strftime("%d %b %Y")}'
    ws['A2'].font = Font(size=10, color='FF64748B')

    row = 4
    indigo_fill = PatternFill('solid', fgColor=indigo)
    light_fill = PatternFill('solid', fgColor=light)
    mid_fill = PatternFill('solid', fgColor=mid)

    s = report.get('sales', {})
    e = report.get('expenses', {})

    # Revenue
    ws.cell(row=row, column=1, value='REVENUE').font = section_font
    row += 1
    for label, col2 in [
        ('Gross Sales', money(s.get('gross_sales'))),
        ('Discounts', -money(s.get('discounts'))),
        ('Net Sales', money(s.get('net_sales'))),
        ('Manual Revenue', money(s.get('manual_revenue'))),
        ('Total Revenue', money(s.get('total_revenue'))),
    ]:
        is_total = 'Total' in label
        row = set_row(ws, row, label, col2,
            font=bold_font if is_total else normal_font,
            fill=mid_fill if is_total else light_fill)

    row += 1
    ws.cell(row=row, column=1, value='COGS & PROFIT').font = section_font
    row += 1
    for label, col2 in [
        ('COGS', money(report.get('cogs'))),
        ('Gross Profit', money(report.get('gross_profit'))),
        ('Gross Margin %', report.get('gross_margin_pct', 0)),
        ('Manual Expenses', money(e.get('manual'))),
        ('Supplier Payments', money(e.get('supplier_payments'))),
        ('Total Expenses', money(e.get('total'))),
        ('Net Profit', money(report.get('net_profit'))),
        ('Net Margin %', report.get('net_margin_pct', 0)),
    ]:
        is_total = 'Net Profit' in label or 'Total' in label
        row = set_row(ws, row, label, col2,
            font=bold_font if is_total else normal_font,
            fill=mid_fill if is_total else light_fill)

    row += 1
    ws.cell(row=row, column=1, value='ACCOUNTS').font = section_font
    row += 1
    for label, col2 in [
        ('Total Collected', money(s.get('total_collected'))),
        ('Accounts Receivable (AR)', money(s.get('accounts_receivable'))),
        ('Supplier Outstanding (AP)', money(e.get('supplier_outstanding'))),
    ]:
        row = set_row(ws, row, label, col2, font=normal_font, fill=light_fill)

    # Trend sheet
    ws2 = wb.create_sheet('Monthly Trend')
    trend_headers = ['Month', 'Revenue', 'Expenses', 'Profit']
    for col, h in enumerate(trend_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = indigo_fill
        c.alignment = Alignment(horizontal='center')

    for i, t in enumerate(report.get('trend', []), 2):
        ws2.cell(row=i, column=1, value=t['month'])
        for j, key in enumerate(['revenue', 'expenses', 'profit'], 2):
            c = ws2.cell(row=i, column=j, value=float(t.get(key, 0)))
            c.number_format = '$#,##0.00'
            c.fill = PatternFill('solid', fgColor=light)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws2.column_dimensions['A'].width = 16
    for col in range(2, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_orders_excel(orders, branch_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'

    indigo = 'FF4F46E5'
    light = 'FFF8FAFC'
    header_font = Font(bold=True, color='FFFFFFFF', size=10)
    indigo_fill = PatternFill('solid', fgColor=indigo)
    light_fill = PatternFill('solid', fgColor=light)
    thin = Side(style='thin', color='FFE2E8F0')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Order #', 'Customer', 'Type', 'Status', 'Payment Status', 'Total', 'Paid', 'Balance', 'Date']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = indigo_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    for i, o in enumerate(orders, 2):
        row_data = [
            o.get('order_number', ''),
            o.get('customer_name') or 'Walk-in',
            o.get('transaction_type', '').replace('_', ' ').title(),
            o.get('status', '').title(),
            o.get('payment_status', '').title(),
            float(o.get('total_amount', 0)),
            float(o.get('amount_paid', 0)),
            float(o.get('balance_due', 0)),
            str(o.get('created_at', ''))[:10],
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = light_fill
            c.border = thin_border
            if col in (6, 7, 8):
                c.number_format = '$#,##0.00'
                c.alignment = Alignment(horizontal='right')

    col_widths = [14, 20, 16, 14, 16, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer